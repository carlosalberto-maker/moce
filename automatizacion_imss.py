import os
import shutil
import unicodedata
import pandas as pd
from playwright.sync_api import sync_playwright
import time
from tkinter import Tk, filedialog

def normalizar_texto(texto):
    """Limpia espacios, convierte a minúsculas y elimina acentos/diacríticos."""
    if pd.isna(texto) or not isinstance(texto, str):
        return ""
    # Convertir a minúsculas y quitar espacios
    t = texto.strip().lower()
    # Eliminar acentos
    t = "".join(c for c in unicodedata.normalize('NFD', t) if unicodedata.category(c) != 'Mn')
    return t

# Mapeo de columna de rol (hoja administrativa) -> (valor_rol, nombre_rol) en MoCE
MAPEO_ROL_COLUMNAS = {
    'asistente medica': (1100, 'Asistente'),
    'asistente medico': (1100, 'Asistente'),
    'administrador de catalogo': (1500, 'Médico'),
    'director': (1500, 'Médico'),
    'jefes de servicios': (1500, 'Médico'),
    'jefe de servicio': (1500, 'Médico'),
    'agenda de citas': (1100, 'Asistente'),
    'agenda citas': (1100, 'Asistente'),
    'enfermera': (1400, 'Enfermera'),
    'enfermero': (1400, 'Enfermera'),
}

def es_hoja_administrativa(hoja):
    """Determina si la hoja corresponde a personal administrativo (u operativo)."""
    h = normalizar_texto(hoja)
    if 'operativ' in h:
        return False
    # 'admin' cubre variantes: 'administrativo', 'adminsitrativo' (typo), 'admin'
    if 'admin' in h:
        return True
    return False

def obtener_columnas_rol(df):
    """Identifica las columnas que marcan el rol en la hoja de personal administrativo."""
    palabras_clave = ['asistente', 'administrador', 'catalogo', 'director', 'jefe', 'agenda', 'enfermera', 'enfermero']
    return [c for c in df.columns.tolist() if any(p in normalizar_texto(str(c)) for p in palabras_clave)]

def detectar_rol_administrativo(row, columnas_rol):
    """Detecta el rol de un usuario administrativo según la columna marcada con 'X'.
    Devuelve (valor_rol, nombre_rol) o None si ningún rol viene marcado."""
    for col in columnas_rol:
        valor = row.get(col)
        if pd.notna(valor) and normalizar_texto(str(valor)) in ('x', 'si', 'true', '1', 's', 'y', 'yes'):
            clave = normalizar_texto(str(col))
            for patron, rol in MAPEO_ROL_COLUMNAS.items():
                if patron in clave:
                    return rol
    return None

def rol_para_usuario(row, es_administrativo, columnas_rol=None):
    """Devuelve (valor_rol, nombre_rol) según el tipo de hoja.
    - Operativo: siempre Médico (1500).
    - Administrativo: detecta la columna marcada con 'X'.
    Devuelve (None, None) si es administrativo y no hay rol marcado."""
    if es_administrativo:
        return detectar_rol_administrativo(row, columnas_rol or [])
    return (1500, 'Médico')

def esperar_alerta(page, texto, timeout=8):
    """Espera a que aparezca una alerta con el texto indicado en el contenedor <app-alert>.
    Devuelve True si la alerta se mostró, False si no apareció dentro del tiempo límite."""
    try:
        page.locator("app-alert .alert-container .alert").filter(has_text=texto).wait_for(
            state="visible", timeout=timeout * 1000
        )
        return True
    except Exception:
        return False

# CONFIGURACIÓN POR DEFECTO
URL_SISTEMA = 'https://ci-moceopd.imss.gob.mx/login'

def buscar_archivo_excel():
    """Abre un diálogo visual para seleccionar el archivo de Excel."""
    print("\n--- SELECCIÓN DE ARCHIVO ---")
    print("Se abrirá una ventana para que selecciones el archivo de Excel...")
    
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    archivo = filedialog.askopenfilename(
        title="Selecciona el archivo de Excel",
        filetypes=[
            ("Archivos de Excel", "*.xlsx"),
            ("Archivos de Excel 97-2003", "*.xls"),
            ("Todos los archivos", "*.*")
        ]
    )
    
    root.destroy()
    
    if not archivo:
        raise FileNotFoundError("No se seleccionó ningún archivo de Excel.")
    
    if not os.path.isfile(archivo):
        raise FileNotFoundError(f"El archivo '{archivo}' no existe.")
    
    print(f"[+] Archivo seleccionado: {os.path.basename(archivo)}")
    print(f"    Ruta: {archivo}")
    return archivo

def seleccionar_hoja(excel_file):
    """Permite seleccionar la hoja del Excel a procesar."""
    xl = pd.ExcelFile(excel_file)
    hojas = xl.sheet_names
    
    if len(hojas) == 1:
        print(f"[+] Procesando la única hoja: '{hojas[0]}'")
        return hojas[0]
        
    print("\n--- SELECCIÓN DE HOJA ---")
    print("El archivo contiene las siguientes hojas:")
    for idx, hoja in enumerate(hojas):
        print(f"  [{idx + 1}] {hoja}")
    while True:
        try:
            opcion = int(input("Selecciona el número de la hoja que deseas procesar: "))
            if 1 <= opcion <= len(hojas):
                return hojas[opcion - 1]
        except ValueError:
            pass
        print("[!] Opción inválida. Intenta de nuevo.")

def obtener_columnas(df):
    """Detecta y confirma las columnas de CURP y Nombre.
    Devuelve (col_curp, col_nombre, idx_curp_excel) donde idx_curp_excel es el índice 1-based de CURP en el Excel."""
    columnas = df.columns.tolist()
    print("\n--- DETECCIÓN DE COLUMNAS ---")
    
    # 1. Detectar CURP
    col_curp = None
    # Prioridad 1: Coincidencia exacta con CURP o Clave Empleado
    exactas_curp = [c for c in columnas if str(c).strip().lower() in ['curp', 'clave empleado', 'clave_empleado', 'clave de empleado', 'claveunica']]
    if exactas_curp:
        default_curp = exactas_curp[0]
    else:
        # Prioridad 2: Que contenga la palabra 'curp'
        contiene_curp = [c for c in columnas if 'curp' in str(c).lower()]
        if contiene_curp:
            default_curp = contiene_curp[0]
        else:
            # Prioridad 3: Que contenga 'clave' o 'empleado', excluyendo cve presupuestal
            contiene_clave = [c for c in columnas if any(x in str(c).lower() for x in ['clave', 'empleado']) and not any(x in str(c).lower() for x in ['presupuestal', 'cve'])]
            if contiene_clave:
                default_curp = contiene_clave[0]
            else:
                # Prioridad 4: Cualquier coincidencia parcial
                posibles_curp = [c for c in columnas if any(x in str(c).lower() for x in ['curp', 'clave', 'empleado'])]
                default_curp = posibles_curp[0] if posibles_curp else (columnas[0] if columnas else '')

    if default_curp:
        respuesta = input(f"¿Usar columna '{default_curp}' para las CURPs? (Presiona ENTER para SÍ, o escribe el nombre correcto): ").strip()
        col_curp = respuesta if respuesta else default_curp
    else:
        while col_curp not in columnas:
            col_curp = input("[!] No se detectó la columna CURP. Escribe el nombre exacto de la columna CURP: ").strip()

    # 2. Detectar Nombre
    col_nombre = None
    # Intentar coincidencia exacta o palabras clave directas
    exactas_nombre = [c for c in columnas if str(c).strip().lower() in ['nombre', 'nombre completo', 'nombres', 'nombre(s)', 'personal', 'médico', 'medico']]
    if exactas_nombre:
        default_nombre = exactas_nombre[0]
    else:
        # Coincidencias parciales excluyendo nombres de unidades/consultorios/etc.
        posibles_nombre = [c for c in columnas if any(x in str(c).lower() for x in ['nombre', 'medico', 'personal', 'completo'])]
        posibles_nombre = [c for c in posibles_nombre if not any(x in str(c).lower() for x in ['unidad', 'consultorio', 'clues', 'hospital', 'clue', 'presupuestal'])]
        default_nombre = posibles_nombre[0] if posibles_nombre else (columnas[0] if columnas else '')

    if default_nombre:
        respuesta = input(f"¿Usar columna '{default_nombre}' para los Nombres? (Presiona ENTER para SÍ, o escribe el nombre correcto): ").strip()
        col_nombre = respuesta if respuesta else default_nombre
    else:
        while col_nombre not in columnas:
            col_nombre = input("[!] No se detectó la columna de Nombre. Escribe el nombre exacto de la columna del Nombre: ").strip()

    # 3. Determinar el índice de la columna CURP en el Excel (1-based para openpyxl)
    idx_curp_excel = columnas.index(col_curp) + 1

    return col_curp, col_nombre, idx_curp_excel

def detectar_columnas_hoja(df):
    """Detecta columnas CURP y Nombre para una hoja dada. Retorna (col_curp, col_nombre)."""
    columnas = df.columns.tolist()

    col_curp = None
    exactas_curp = [c for c in columnas if str(c).strip().lower() in ['curp', 'clave empleado', 'clave_empleado', 'clave de empleado', 'claveunica']]
    if exactas_curp:
        col_curp = exactas_curp[0]
    else:
        contiene_curp = [c for c in columnas if 'curp' in str(c).lower()]
        if contiene_curp:
            col_curp = contiene_curp[0]
        else:
            contiene_clave = [c for c in columnas if any(x in str(c).lower() for x in ['clave', 'empleado']) and not any(x in str(c).lower() for x in ['presupuestal', 'cve'])]
            if contiene_clave:
                col_curp = contiene_clave[0]
            else:
                posibles_curp = [c for c in columnas if any(x in str(c).lower() for x in ['curp', 'clave', 'empleado'])]
                col_curp = posibles_curp[0] if posibles_curp else (columnas[0] if columnas else '')

    col_nombre = None
    exactas_nombre = [c for c in columnas if str(c).strip().lower() in ['nombre', 'nombre completo', 'nombres', 'nombre(s)', 'personal', 'médico', 'medico']]
    if exactas_nombre:
        col_nombre = exactas_nombre[0]
    else:
        posibles_nombre = [c for c in columnas if any(x in str(c).lower() for x in ['nombre', 'medico', 'personal', 'completo'])]
        posibles_nombre = [c for c in posibles_nombre if not any(x in str(c).lower() for x in ['unidad', 'consultorio', 'clues', 'hospital', 'clue', 'presupuestal'])]
        col_nombre = posibles_nombre[0] if posibles_nombre else (columnas[0] if columnas else '')

    return col_curp, col_nombre


def ejecutar_automatizacion():
    print("=" * 60)
    print("        ASISTENTE DE AUTOMATIZACIÓN - IMSS MOCE")
    print("=" * 60)
    
    try:
        excel_file = buscar_archivo_excel()
        xl = pd.ExcelFile(excel_file)
        todas_las_hojas = xl.sheet_names
        print(f"\n[+] Hojas encontradas en el archivo: {todas_las_hojas}")

        col_matricula = "Matricula"
        datos_por_hoja = {}

        for hoja in todas_las_hojas:
            print(f"\n{'─' * 50}")
            print(f"[+] Procesando hoja: '{hoja}'")
            print(f"{'─' * 50}")

            df = pd.read_excel(excel_file, sheet_name=hoja)

            columnas_rol = obtener_columnas_rol(df)
            es_administrativo = es_hoja_administrativa(hoja) or bool(columnas_rol)
            if es_administrativo:
                print(f"    [+] Detectada como PERSONAL ADMINISTRATIVO. Columnas de rol: {columnas_rol}")
            else:
                print("    [+] Detectada como PERSONAL OPERATIVO. Rol por defecto: Médico.")

            col_curp, col_nombre = detectar_columnas_hoja(df)
            if not col_curp or not col_nombre:
                print(f"    [!] No se detectaron columnas CURP/Nombre en la hoja '{hoja}'. Se omitirá.")
                continue

            if col_matricula not in df.columns:
                idx_col_curp = df.columns.get_loc(col_curp)
                df.insert(idx_col_curp + 1, col_matricula, "")
                print(f"    [+] Columna '{col_matricula}' insertada junto a '{col_curp}'.")
            else:
                print(f"    [+] Columna '{col_matricula}' ya existente.")
            df[col_matricula] = df[col_matricula].astype(object)

            df_medicos = df[df[col_curp].notna()].copy()
            filtro_vacias = df_medicos[col_matricula].isna() | (df_medicos[col_matricula].astype(str).str.strip() == "") | (df_medicos[col_matricula].astype(str).str.contains("ERROR|NO GENERADA", case=False, na=True))
            df_pendientes = df_medicos[filtro_vacias].copy()
            cant_omitidos = len(df_medicos) - len(df_pendientes)
            if cant_omitidos > 0:
                print(f"    [i] Se omitirán {cant_omitidos} registros ya procesados.")

            print(f"    [i] Registros pendientes en '{hoja}': {len(df_pendientes)}")
            datos_por_hoja[hoja] = {
                'df': df,
                'col_curp': col_curp,
                'col_nombre': col_nombre,
                'col_matricula': col_matricula,
                'es_administrativo': es_administrativo,
                'columnas_rol': columnas_rol,
                'df_pendientes': df_pendientes,
            }

    except Exception as e:
        print(f"\n[!] Error inicial: {e}")
        return

    total_pendientes = sum(len(d['df_pendientes']) for d in datos_por_hoja.values())
    print(f"\n>>> Total de registros pendientes en TODAS las hojas: {total_pendientes}")
    if total_pendientes == 0:
        print("[!] No hay registros nuevos para procesar.")
        return

    with sync_playwright() as p:
        # Intentar abrir Edge (o Chromium si Edge no está instalado)
        print("\n>>> Iniciando navegador...")
        try:
            browser = p.chromium.launch(headless=False, channel="msedge")
        except Exception as e:
            print(f"[i] No se pudo abrir Edge mediante canal 'msedge' ({e}). Iniciando Chromium normal...")
            browser = p.chromium.launch(headless=False)
            
        context = browser.new_context()
        page = context.new_page()

        print(f">>> Navegando a {URL_SISTEMA}...")
        page.goto(URL_SISTEMA)

        print("\n" + "!" * 50)
        print("--- ACCIÓN MANUAL REQUERIDA ---")
        print("1. Inicia sesión en la ventana de Edge que se abrió.")
        print("2. Navega hasta estar dentro de la pestaña 'Alta de Usuarios'.")
        print("3. Cuando ya veas los campos de CURP y Matrícula listos:")
        print("   Regresa a esta ventana negra y presiona ENTER.")
        print("!" * 50 + "\n")
        
        input("Presiona ENTER aquí cuando ya estés en la pestaña de Alta de Usuarios...")

        exitosos = 0
        errores = 0
        contador_global = 0
        curps_exitosas = set()
        curp_matricula_map = {}

        for hoja, datos in datos_por_hoja.items():
            df = datos['df']
            col_curp = datos['col_curp']
            col_nombre = datos['col_nombre']
            col_matricula = datos['col_matricula']
            es_administrativo = datos['es_administrativo']
            columnas_rol = datos['columnas_rol']
            df_pendientes = datos['df_pendientes']

            print(f"\n{'═' * 50}")
            print(f">>> PROCESANDO HOJA: '{hoja}' ({len(df_pendientes)} pendientes)")
            print(f"{'═' * 50}")

            for index, row in df_pendientes.iterrows():
                contador_global += 1
                curp_raw = str(row[col_curp]).strip()
                if curp_raw.endswith('.0'):
                    curp = curp_raw[:-2]
                else:
                    curp = curp_raw
                    
                nombre = str(row[col_nombre]).strip()
                print(f"\n[*] [{contador_global}/{total_pendientes}] Hoja '{hoja}' | {nombre} | CURP: {curp}")

                if curp in curps_exitosas:
                    matricula_conocida = curp_matricula_map[curp]
                    df.at[index, col_matricula] = matricula_conocida
                    print(f"    [i] CURP duplicada ya procesada. Matrícula aplicada: {matricula_conocida}")
                    exitosos += 1
                    continue

                # Determinar rol según el tipo de hoja
                rol_det = rol_para_usuario(row, es_administrativo, columnas_rol)
                if rol_det is None:
                    print("    [!] No se detectó un rol marcado (X) para este usuario. Se omitirá.")
                    df.at[index, col_matricula] = "ERROR: Sin rol marcado"
                    errores += 1
                    continue
                rol_value, rol_nombre = rol_det
                print(f"    [i] Rol asignado: {rol_nombre} ({rol_value})")

                try:
                    # 1. Ingresar CURP de manera robusta
                    campo_curp = None
                    selectores_curp = [
                        "input[formcontrolname='curp']",
                        "input:near(label:text('CURP'))",
                        "input[placeholder*='CURP' i]",
                        "input[id*='curp' i]",
                        "input[name*='curp' i]",
                        "input[type='text']"
                    ]
                    
                    for selector in selectores_curp:
                        try:
                            locator = page.locator(selector).first
                            if locator.is_visible(timeout=2000):
                                campo_curp = locator
                                break
                        except Exception:
                            continue
                    
                    if campo_curp is None:
                        campo_curp = page.locator("input[formcontrolname='curp']").first
                    
                    campo_curp.click()
                    campo_curp.press("Control+A")
                    campo_curp.press("Backspace")
                    campo_curp.fill(curp)
                    
                    # 2. Clic en Búsqueda de CURP
                    btn_buscar = None
                    selectores_buscar = [
                        "button:has-text('Búsqueda de CURP')",
                        "button:has(.bi-search)",
                        "button.btn-success",
                        "button:has-text('Buscar')"
                    ]
                    for selector in selectores_buscar:
                        try:
                            locator = page.locator(selector).first
                            if locator.is_visible(timeout=2000):
                                btn_buscar = locator
                                break
                        except Exception:
                            continue
                    
                    if btn_buscar is None:
                        btn_buscar = page.locator("button:has-text('Búsqueda de CURP')").first
                    
                    btn_buscar.click()
                    time.sleep(2.5) 
                    
                    # 3. Capturar Matrícula
                    campo_matricula = None
                    selectores_matricula = [
                        "input[formcontrolname='matricula']",
                        "input:near(label:text('Matrícula'))",
                        "input:near(label:text('Matricula'))",
                        "input[placeholder*='Matrícula' i]"
                    ]
                    
                    for selector in selectores_matricula:
                        try:
                            locator = page.locator(selector).first
                            if locator.is_visible(timeout=2000):
                                campo_matricula = locator
                                break
                        except Exception:
                            continue
                    
                    if campo_matricula is None:
                        campo_matricula = page.locator("input[formcontrolname='matricula']").first
                    
                    matricula = campo_matricula.input_value()
                    
                    if matricula and matricula.strip():
                        print(f"    [+] Matrícula obtenida con éxito: {matricula}")
                        df.at[index, col_matricula] = matricula
                        curps_exitosas.add(curp)
                        curp_matricula_map[curp] = matricula
                        exitosos += 1
                        
                        # 4. Seleccionar Rol
                        try:
                            try:
                                page.select_option("select[formcontrolname='rol']", value=str(rol_value))
                            except Exception:
                                page.select_option("select:has-text('Seleccionar rol')", label=rol_nombre)
                            print(f"    [+] Rol seleccionado: {rol_nombre} ({rol_value})")
                        except Exception as e_rol:
                            print(f"    [i] Advertencia al seleccionar rol: {e_rol}")

                        # 5. Clic en Agregar
                        btn_agregar = None
                        selectores_agregar = [
                            "button:has-text('Agregar')",
                            "button:has(.bi-plus-circle)",
                            "button.btn-secondary:has-text('Agregar')"
                        ]
                        for selector in selectores_agregar:
                            try:
                                locator = page.locator(selector).first
                                if locator.is_visible(timeout=2000):
                                    btn_agregar = locator
                                    break
                            except Exception:
                                continue
                        
                        if btn_agregar is None:
                            btn_agregar = page.locator("button:has-text('Agregar')").first
                        
                        try:
                            if btn_agregar.is_disabled(timeout=1500):
                                print("    [!] El botón 'Agregar' está deshabilitado en el sistema. Omitiendo clic.")
                                df.at[index, col_matricula] = "ERROR: Botón Agregar deshabilitado"
                                errores += 1
                            else:
                                btn_agregar.click(timeout=2000)
                                print("    [+] Clic en Agregar ejecutado.")
                                
                                agregado_ok = esperar_alerta(page, "Usuario agregado correctamente")
                                if agregado_ok:
                                    print("    [+] Confirmación detectada: 'Usuario agregado correctamente'.")
                                else:
                                    print("    [!] NO se detectó la confirmación 'Usuario agregado correctamente'.")
                                time.sleep(1.5)
                                
                                # 5.1 Clic en Continuar
                                btn_continuar = None
                                selectores_continuar = [
                                    "button:has-text('Continuar')",
                                    "button:has(.bi-arrow-right-circle)",
                                    "button.btn-primary:has-text('Continuar')"
                                ]
                                for selector in selectores_continuar:
                                    try:
                                        locator = page.locator(selector).first
                                        if locator.is_visible(timeout=2000):
                                            btn_continuar = locator
                                            break
                                    except Exception:
                                        continue
                                
                                if btn_continuar is None:
                                    btn_continuar = page.locator("button:has-text('Continuar')").first
                                    
                                try:
                                    btn_continuar.click(timeout=2000)
                                    print("    [+] Clic en Continuar ejecutado.")
                                    
                                    guardado_ok = esperar_alerta(page, "guardado exitosamente")
                                    if guardado_ok:
                                        print("    [+] Confirmación detectada: 'guardado exitosamente'.")
                                    else:
                                        print("    [!] NO se detectó la confirmación 'guardado exitosamente'.")
                                    time.sleep(2)
                                except Exception as e_cont:
                                    print(f"    [!] Advertencia al hacer clic en Continuar: {e_cont}")
                                    time.sleep(1)
                                    
                                # 5.2 Regresar a Alta de Usuarios
                                print("    [i] Regresando a la pestaña 'Alta de Usuarios'...")
                                regreso_exitoso = False
                                selectores_regreso = [
                                    "button[ngbnavlink]:has-text('Alta de Usuarios')",
                                    "button:has-text('Alta de Usuarios')",
                                    "#ngb-nav-1",
                                    "a:has-text('Alta de Usuarios')",
                                    "span:has-text('Alta de Usuarios')",
                                    "text=Alta de Usuarios"
                                ]
                                for selector in selectores_regreso:
                                    try:
                                        locator = page.locator(selector).first
                                        if locator.is_visible(timeout=2000):
                                            locator.click()
                                            regreso_exitoso = True
                                            break
                                    except Exception:
                                        continue
                                
                                if not regreso_exitoso:
                                    print("    [!] No se pudo regresar automáticamente a 'Alta de Usuarios'.")
                                    try:
                                        page.locator("input[formcontrolname='curp']").wait_for(state="visible", timeout=8000)
                                        print("    [+] Detectada pantalla de Alta de Usuarios de nuevo.")
                                    except Exception:
                                        input("    Presiona ENTER cuando ya estés de vuelta en 'Alta de Usuarios'...")
                                else:
                                    try:
                                        page.locator("input[formcontrolname='curp']").wait_for(state="visible", timeout=5000)
                                        time.sleep(1.5)
                                    except Exception:
                                        time.sleep(2)

                                # 5.3 Eliminar el registro anterior
                                try:
                                    btn_trash = None
                                    selectores_trash = [
                                        "button.btn-danger:has(.bi-trash)",
                                        "button:has(.bi-trash)",
                                        "i.bi-trash"
                                    ]
                                    for selector in selectores_trash:
                                        try:
                                            locator = page.locator(selector).first
                                            if locator.is_visible(timeout=1500):
                                                btn_trash = locator
                                                break
                                        except Exception:
                                            continue
                                    if btn_trash is not None:
                                        btn_trash.click(timeout=2000)
                                        print("    [+] Registro anterior eliminado.")
                                        time.sleep(1.5)
                                    else:
                                        print("    [i] No hay registro anterior para eliminar.")
                                except Exception as e_trash:
                                    print(f"    [i] Advertencia al eliminar registro anterior: {e_trash}")
                        except Exception as e_btn:
                            print(f"    [!] Error al interactuar con el botón Agregar: {e_btn}")
                            df.at[index, col_matricula] = "ERROR: Falla botón Agregar"
                            errores += 1
                    else:
                        print("    [!] El sistema no mostró matrícula para esta CURP.")
                        df.at[index, col_matricula] = "NO GENERADA"
                        errores += 1

                except Exception as e:
                    print(f"    [!] Error al procesar a este usuario: {e}")
                    df.at[index, col_matricula] = f"ERROR: {str(e)[:50]}"
                    errores += 1
                    continue

        curps_fallidas_a_reintentar = []
        for hoja_r, datos_r in datos_por_hoja.items():
            df_r = datos_r['df']
            col_curp_r = datos_r['col_curp']
            col_matricula_r = datos_r['col_matricula']
            for idx_r in df_r.index:
                curp_r_raw = str(df_r.at[idx_r, col_curp_r]).strip() if pd.notna(df_r.at[idx_r, col_curp_r]) else ""
                if curp_r_raw.endswith('.0'):
                    curp_r_raw = curp_r_raw[:-2]
                if not curp_r_raw or curp_r_raw in curps_exitosas:
                    continue
                mat_r = df_r.at[idx_r, col_matricula_r]
                es_fallo = (pd.isna(mat_r) or str(mat_r).strip() == "" or
                           "ERROR" in str(mat_r).upper() or "NO GENERADA" in str(mat_r).upper())
                if es_fallo:
                    curps_fallidas_a_reintentar.append((hoja_r, idx_r, curp_r_raw))

        if curps_fallidas_a_reintentar:
            print(f"\n{'═' * 50}")
            print(f">>> SEGUNDA VUELTA: {len(curps_fallidas_a_reintentar)} CURPs fallidas para reintentar")
            print(f"{'═' * 50}")

            selectores_curp_r = [
                "input[formcontrolname='curp']",
                "input:near(label:text('CURP'))",
                "input[placeholder*='CURP' i]",
                "input[id*='curp' i]",
                "input[name*='curp' i]",
                "input[type='text']"
            ]
            selectores_buscar_r = [
                "button:has-text('Búsqueda de CURP')",
                "button:has(.bi-search)",
                "button.btn-success",
                "button:has-text('Buscar')"
            ]
            selectores_matricula_r = [
                "input[formcontrolname='matricula']",
                "input:near(label:text('Matrícula'))",
                "input:near(label:text('Matricula'))",
                "input[placeholder*='Matrícula' i]"
            ]
            selectores_agregar_r = [
                "button:has-text('Agregar')",
                "button:has(.bi-plus-circle)",
                "button.btn-secondary:has-text('Agregar')"
            ]
            selectores_continuar_r = [
                "button:has-text('Continuar')",
                "button:has(.bi-arrow-right-circle)",
                "button.btn-primary:has-text('Continuar')"
            ]
            selectores_regreso_r = [
                "button[ngbnavlink]:has-text('Alta de Usuarios')",
                "button:has-text('Alta de Usuarios')",
                "#ngb-nav-1",
                "a:has-text('Alta de Usuarios')",
                "span:has-text('Alta de Usuarios')",
                "text=Alta de Usuarios"
            ]
            selectores_trash_r = [
                "button.btn-danger:has(.bi-trash)",
                "button:has(.bi-trash)",
                "i.bi-trash"
            ]

            for hoja_r, idx_r, curp_r in curps_fallidas_a_reintentar:
                contador_global += 1
                datos_r = datos_por_hoja[hoja_r]
                df_r = datos_r['df']
                col_curp_r = datos_r['col_curp']
                col_nombre_r = datos_r['col_nombre']
                col_matricula_r = datos_r['col_matricula']
                es_admin_r = datos_r['es_administrativo']
                cols_rol_r = datos_r['columnas_rol']

                nombre_r = str(df_r.at[idx_r, col_nombre_r]).strip()
                print(f"\n[*] [REINTENTO {contador_global}] Hoja '{hoja_r}' | {nombre_r} | CURP: {curp_r}")

                rol_det_r = rol_para_usuario(df_r.loc[idx_r], es_admin_r, cols_rol_r)
                if rol_det_r is None:
                    print("    [!] No se detectó un rol marcado (X) para este usuario. Se omitirá.")
                    df_r.at[idx_r, col_matricula_r] = "ERROR: Sin rol marcado"
                    errores += 1
                    continue
                rol_value_r, rol_nombre_r = rol_det_r
                print(f"    [i] Rol asignado: {rol_nombre_r} ({rol_value_r})")

                try:
                    campo_curp_r = None
                    for selector in selectores_curp_r:
                        try:
                            locator = page.locator(selector).first
                            if locator.is_visible(timeout=2000):
                                campo_curp_r = locator
                                break
                        except Exception:
                            continue
                    if campo_curp_r is None:
                        campo_curp_r = page.locator("input[formcontrolname='curp']").first
                    campo_curp_r.click()
                    campo_curp_r.press("Control+A")
                    campo_curp_r.press("Backspace")
                    campo_curp_r.fill(curp_r)

                    btn_buscar_r = None
                    for selector in selectores_buscar_r:
                        try:
                            locator = page.locator(selector).first
                            if locator.is_visible(timeout=2000):
                                btn_buscar_r = locator
                                break
                        except Exception:
                            continue
                    if btn_buscar_r is None:
                        btn_buscar_r = page.locator("button:has-text('Búsqueda de CURP')").first
                    btn_buscar_r.click()
                    time.sleep(2.5)

                    campo_mat_r = None
                    for selector in selectores_matricula_r:
                        try:
                            locator = page.locator(selector).first
                            if locator.is_visible(timeout=2000):
                                campo_mat_r = locator
                                break
                        except Exception:
                            continue
                    if campo_mat_r is None:
                        campo_mat_r = page.locator("input[formcontrolname='matricula']").first
                    matricula_r = campo_mat_r.input_value()

                    if matricula_r and matricula_r.strip():
                        print(f"    [+] Matrícula obtenida en reintento: {matricula_r}")
                        df_r.at[idx_r, col_matricula_r] = matricula_r
                        curps_exitosas.add(curp_r)
                        curp_matricula_map[curp_r] = matricula_r
                        exitosos += 1

                        for other_idx in df_r.index:
                            if other_idx != idx_r:
                                other_curp = str(df_r.at[other_idx, col_curp_r]).strip() if pd.notna(df_r.at[other_idx, col_curp_r]) else ""
                                if other_curp.endswith('.0'):
                                    other_curp = other_curp[:-2]
                                if other_curp == curp_r:
                                    other_mat = df_r.at[other_idx, col_matricula_r]
                                    es_fallo_o = (pd.isna(other_mat) or str(other_mat).strip() == "" or
                                                 "ERROR" in str(other_mat).upper() or "NO GENERADA" in str(other_mat).upper())
                                    if es_fallo_o:
                                        df_r.at[other_idx, col_matricula_r] = matricula_r
                                        print(f"    [+] Matrícula también aplicada a fila duplicada (índice {other_idx})")

                        try:
                            try:
                                page.select_option("select[formcontrolname='rol']", value=str(rol_value_r))
                            except Exception:
                                page.select_option("select:has-text('Seleccionar rol')", label=rol_nombre_r)
                            print(f"    [+] Rol seleccionado: {rol_nombre_r} ({rol_value_r})")
                        except Exception as e_rol:
                            print(f"    [i] Advertencia al seleccionar rol: {e_rol}")

                        btn_agregar_r = None
                        for selector in selectores_agregar_r:
                            try:
                                locator = page.locator(selector).first
                                if locator.is_visible(timeout=2000):
                                    btn_agregar_r = locator
                                    break
                            except Exception:
                                continue
                        if btn_agregar_r is None:
                            btn_agregar_r = page.locator("button:has-text('Agregar')").first

                        try:
                            if btn_agregar_r.is_disabled(timeout=1500):
                                print("    [!] El botón 'Agregar' está deshabilitado. Omitiendo clic.")
                                df_r.at[idx_r, col_matricula_r] = "ERROR: Botón Agregar deshabilitado"
                                errores += 1
                            else:
                                btn_agregar_r.click(timeout=2000)
                                print("    [+] Clic en Agregar ejecutado.")

                                agregado_ok = esperar_alerta(page, "Usuario agregado correctamente")
                                if agregado_ok:
                                    print("    [+] Confirmación: 'Usuario agregado correctamente'.")
                                else:
                                    print("    [!] NO se detectó confirmación 'Usuario agregado correctamente'.")
                                time.sleep(1.5)

                                btn_continuar_r = None
                                for selector in selectores_continuar_r:
                                    try:
                                        locator = page.locator(selector).first
                                        if locator.is_visible(timeout=2000):
                                            btn_continuar_r = locator
                                            break
                                    except Exception:
                                        continue
                                if btn_continuar_r is None:
                                    btn_continuar_r = page.locator("button:has-text('Continuar')").first

                                try:
                                    btn_continuar_r.click(timeout=2000)
                                    print("    [+] Clic en Continuar ejecutado.")
                                    guardado_ok = esperar_alerta(page, "guardado exitosamente")
                                    if guardado_ok:
                                        print("    [+] Confirmación: 'guardado exitosamente'.")
                                    else:
                                        print("    [!] NO se detectó confirmación 'guardado exitosamente'.")
                                    time.sleep(2)
                                except Exception as e_cont:
                                    print(f"    [!] Advertencia al hacer clic en Continuar: {e_cont}")
                                    time.sleep(1)

                                print("    [i] Regresando a 'Alta de Usuarios'...")
                                regreso_ok = False
                                for selector in selectores_regreso_r:
                                    try:
                                        locator = page.locator(selector).first
                                        if locator.is_visible(timeout=2000):
                                            locator.click()
                                            regreso_ok = True
                                            break
                                    except Exception:
                                        continue
                                if not regreso_ok:
                                    print("    [!] No se pudo regresar automáticamente a 'Alta de Usuarios'.")
                                    try:
                                        page.locator("input[formcontrolname='curp']").wait_for(state="visible", timeout=8000)
                                        print("    [+] Detectada pantalla de Alta de Usuarios.")
                                    except Exception:
                                        input("    Presiona ENTER cuando estés en 'Alta de Usuarios'...")
                                else:
                                    try:
                                        page.locator("input[formcontrolname='curp']").wait_for(state="visible", timeout=5000)
                                        time.sleep(1.5)
                                    except Exception:
                                        time.sleep(2)

                                try:
                                    btn_trash_r = None
                                    for selector in selectores_trash_r:
                                        try:
                                            locator = page.locator(selector).first
                                            if locator.is_visible(timeout=1500):
                                                btn_trash_r = locator
                                                break
                                        except Exception:
                                            continue
                                    if btn_trash_r is not None:
                                        btn_trash_r.click(timeout=2000)
                                        print("    [+] Registro anterior eliminado.")
                                        time.sleep(1.5)
                                    else:
                                        print("    [i] No hay registro anterior para eliminar.")
                                except Exception as e_trash:
                                    print(f"    [i] Advertencia al eliminar registro anterior: {e_trash}")
                        except Exception as e_btn:
                            print(f"    [!] Error al interactuar con el botón Agregar: {e_btn}")
                            df_r.at[idx_r, col_matricula_r] = "ERROR: Falla botón Agregar"
                            errores += 1
                    else:
                        print("    [!] El sistema no mostró matrícula para esta CURP (reintento).")
                        df_r.at[idx_r, col_matricula_r] = "NO GENERADA"
                        errores += 1

                except Exception as e:
                    print(f"    [!] Error al procesar (reintento): {e}")
                    df_r.at[idx_r, col_matricula_r] = f"ERROR: {str(e)[:50]}"
                    errores += 1
                    continue
        else:
            print("\n[i] No hay CURPs fallidas pendientes para una segunda vuelta.")

        # 6. Guardar resultados en el mismo archivo de origen
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            base, ext = os.path.splitext(excel_file)
            backup_file = f"{base}_respaldo{ext}"
            shutil.copy2(excel_file, backup_file)

            wb = load_workbook(excel_file)
            red_fill = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid")

            for hoja, datos in datos_por_hoja.items():
                df = datos['df']
                col_curp = datos['col_curp']
                col_matricula = datos['col_matricula']

                if hoja not in wb.sheetnames:
                    print(f"[!] La hoja '{hoja}' no existe en el workbook. Se omitirá el guardado.")
                    continue

                ws = wb[hoja]

                encabezados = {}
                for cell in ws[1]:
                    if cell.value:
                        encabezados[str(cell.value).strip().lower()] = cell.column

                col_curp_lower = str(col_curp).strip().lower()
                col_curp_idx = None
                for key, val in encabezados.items():
                    if col_curp_lower in key or key in col_curp_lower:
                        col_curp_idx = val
                        break

                if col_curp_idx is None:
                    columnas_list = df.columns.tolist()
                    col_curp_idx = columnas_list.index(col_curp) + 1

                col_matricula_idx = col_curp_idx + 1

                col_matricula_excel = None
                for cell in ws[1]:
                    if cell.value and 'matri' in str(cell.value).strip().lower():
                        col_matricula_excel = cell.column
                        break

                if col_matricula_excel is None:
                    ws.insert_cols(col_matricula_idx)
                    ws.cell(row=1, column=col_matricula_idx, value="Matricula")
                    print(f"[+] Hoja '{hoja}': Columna 'Matricula' insertada en columna {chr(64 + col_matricula_idx) if col_matricula_idx <= 26 else col_matricula_idx}.")
                else:
                    col_matricula_idx = col_matricula_excel

                curp_col_in_df = df.columns.get_loc(col_curp)
                mat_col_in_df = df.columns.get_loc(col_matricula)

                mapa_filas = {}
                for fila_excel in range(2, ws.max_row + 1):
                    curp_excel = ws.cell(row=fila_excel, column=col_curp_idx).value
                    if curp_excel:
                        curp_limpia = str(curp_excel).strip()
                        if curp_limpia.endswith('.0'):
                            curp_limpia = curp_limpia[:-2]
                        if curp_limpia not in mapa_filas:
                            mapa_filas[curp_limpia] = []
                        mapa_filas[curp_limpia].append(fila_excel)

                for _, row in df.iterrows():
                    curp_val = str(row[col_curp]).strip()
                    if curp_val.endswith('.0'):
                        curp_val = curp_val[:-2]
                    mat_val = row[col_matricula]

                    if curp_val in mapa_filas:
                        for fila_excel in mapa_filas[curp_val]:
                            ws.cell(row=fila_excel, column=col_matricula_idx, value=str(mat_val) if mat_val else "")

                for fila_excel in range(2, ws.max_row + 1):
                    valor = ws.cell(row=fila_excel, column=col_matricula_idx).value
                    es_error = (valor is None or not str(valor).strip()
                                or "ERROR" in str(valor).upper()
                                or "NO GENERADA" in str(valor).upper())
                    if es_error:
                        ws.cell(row=fila_excel, column=col_matricula_idx).fill = red_fill
                        ws.cell(row=fila_excel, column=col_curp_idx).fill = red_fill

            wb.save(excel_file)
            print(f"\n>>> Archivo actualizado: {os.path.basename(excel_file)}")

            print("\n" + "=" * 60)
            print(f">>> PROCESO FINALIZADO <<<")
            print(f" - Hojas procesadas: {len(datos_por_hoja)}")
            print(f" - Registros procesados con éxito: {exitosos}")
            print(f" - Registros con error o vacíos: {errores}")
            print(f" - Archivo guardado en: {excel_file}")
            print(f" - Respaldo en: {backup_file}")
            print("=" * 60)
        except Exception as e:
            print(f"\n[!] Error al escribir el archivo: {e}")
            print("[i] El archivo original NO fue modificado. El respaldo está intacto.")
        
        input("\nPresiona ENTER aquí para cerrar el navegador y salir...")
        browser.close()

if __name__ == "__main__":
    ejecutar_automatizacion()
